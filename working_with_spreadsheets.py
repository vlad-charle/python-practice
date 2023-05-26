import openpyxl

# open xlsx
employees_file = openpyxl.load_workbook("employees.xlsx")
employees_list = employees_file["Sheet1"]

# create new xlsx
updated_employee_file = openpyxl.Workbook()
updated_employees_list = updated_employee_file.active

# populate header 
updated_employees_list.append(["Name", "Years of Experience"])

# read data in xlsx and create dict for sorting
employees_dict = {}
for employee_row in range(2, employees_list.max_row + 1):

    employee_name = employees_list.cell(employee_row, 1).value

    employee_experience = int(employees_list.cell(employee_row, 2).value)
    employees_dict[employee_name] = employee_experience

# sort desc
employees_dict_sorted = dict(sorted(employees_dict.items(), key=lambda item: item[1], reverse=True))

# write data to new xlsx from sorted dict
index = 2 # starting row, 1st row is header
for employee in employees_dict_sorted:

    updated_employee_name = updated_employees_list.cell(index, 1)
    updated_employee_name.value = employee

    updated_employee_experience = updated_employees_list.cell(index, 2)
    updated_employee_experience.value = employees_dict_sorted[employee]

    index += 1
    
# save new xlsx
updated_employee_file.save("employees_updated.xlsx")